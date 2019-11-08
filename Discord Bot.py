import discord
import datetime
import openpyxl
import asyncio
from json import loads
import configparser
import random

client = discord.Client()


@client.event
async def on_ready():
    print(client.user.id)
    print("ready")
    game = discord.Game('7개의 서버와 함께 하는중')
    await client.change_presence(status=discord.Status.online, activity=game)


@client.event
async def on_message(message):
    if message.content.startswith("!정보"):
        date = datetime.datetime.utcfromtimestamp(((int(message.author.id) >> 22) + 1420070400000) / 1000)
        embed = discord.Embed(color=0x00ff00)
        embed.add_field(name="이름", value=message.author.name, inline=True)
        embed.add_field(name="서버 닉네임", value=message.author.display_name, inline=True)
        embed.add_field(name="가입일", value=str(date.year) + "년" + str(date.month) + "월" + str(date.day) + "일", inline=True)
        embed.add_field(name="아이디", value=message.author.id, inline=True)
        embed.set_thumbnail(url=message.author.avatar_url)
        await message.channel.send(embed=embed)
    if message.content.startswith("!안녕"):
        await message.channel.send("안녕하세요 전 Kidong Bot 입니다!")
    if message.content.startswith("!잘가"):
        await message.channel.send("야호! 퇴근이다! @===^--^")
    if message.content.startswith("!차량목록"):
        await message.channel.send("Bugatti chiron 3,113,712,000원\nBugatti veyron 2,983,974,000원\nBugatti Divo 6,486,900,000원\nBugatti La Voiture Noire 142,000,000,000원\nFerarri ltalia 490,000.000원\nLamborgini Terzo 120,000,000,000원 이상")
    if message.content.startswith("!오목"):
        await message.channel.send("전 오목을 할줄 몰라요. ~~개발자가 놀고 있어서 오목 시스템이 없어요.~~")
    if message.content.startswith("!문의"):
        await message.channel.send("문의는 @Kidong 에게 해주시면 됩니다.")
    if message.content.startswith("!on"):
        await message.channel.send("@everyone\nKidong 이 다시 돌아왔습니다!\n문의는 @kidong#3103 에게 부탁드립니다.")
    if message.content.startswith("!off"):
        await message.channel.send("@everyone\nKidong이 잠시 자리를 비웠습니다.\n문의는 DM 주시면 온라인일때 처리 해드리겠습니다.")
    if message.content.startswith("!봇제작자"):
        await message.channel.send("봇 제작자는 Kidong 이며, 1인 개발자 입니다.\nKidong 봇은 개발자가 초대한 서버에만 사용할수 있으며\n그 외 개발자가 초대하지 않은 서버는 사용을 하지 말아주시길 바랍니다.")
    if message.content.startswith("!wshutdown"):
        await message.channel.send("✔셧다운 합니다✔")
    if message.content.startswith("!Urgent shutdown"):
        await message.channel.send("✔긴급 셧다운 되었습니다.✔")
    if message.content.startswith("!디스코드링크"):
        await message.channel.send("https://discord.gg/q8UQca5")
    if message.content.startswith("!이의제기서류"):
        await message.channel.send("[이의제기 서류]\n작성자는 자신의 처벌이 너무 불공평하다, 등\n여러 사유로 서류를 작성하실수 있으며, 이 이의제기 서류를 통해서 관리자에게 이의를 제기 할수 있습니다.\n작성자는 이의 제기를 하더라도 처벌의 강도가 달라지지 않을수도 있다는 점\n등 이부분에 대해서는 명심하시길 바랍니다.\n------------------------------------------\n ")
    if message.content.startswith("!GTA상품"):
        await message.channel.send("```\n[GTA5 상품 목록]\nGTA5 계정 (스팀) 20,000원\n------------------------\nKingpin Hack 40.000원\nPostion Hack 30,000원\n레퀴엠 [ VIP 영구적으로 사용할수 있게 해드립니다. ] 70,000원\nKidong Hack [ 업데이트 되면 더욱 비싸집니다. ] 25,000원\nGTA5 제품 키 25,000원\n모든 상품은 문화상품권으로만 결제가 가능합니다.\n------------------------\n[ GTA5 Hack 대리 ]\n초급 패키지\n돈 5억 + 벙커 언락 + 레벨 1 ~ 300\n가격 : 5,000원\n중급 패키지\n돈 8억 + 벙커 언락 + 레벨 1 ~ 8000\n가격 : 10,000\n고급 패키지\n돈 11억 + 벙커 언락 + 레벨  1 ~ 8000\n가격 : 15,000원\n------------------------\n구매문의는 @Kidong#3103\n```")
    if message.content.startswith("!서류"):
        await message.channel.send("[ 서류 작성 ]\n구매하고 싶은 상품:\n문화 상품권(핀번호):\n{주의 사항}\n판매자는 핵 판매 이후 일어나는 일들,\n[#예: 밴, 영구정지, 등] 사유에 대해 책임을 지지 않습니다.\n[대리도 저 조건에 포함]\n만약 이 조건을 지키지 못하시겠다면 구매를 취소해주시길 바랍니다.\n서명:\n구매 날짜: 0000년 0월 00일")
    if message.content.startswith("!후원문의"):
        await message.channel.send("후원 문의는 총관리자분들에게 해주시면 됩니다.")
    if message.content.startswith("!후원방법"):
        await message.channel.send("[후원 서류 작성]\n후원 할 금액:\n후원 방법(문화상품권 이면 핀번호 적어주세요):\n{주의 사항}\n후원 하신 금액은 오류 등 이유로 지급을 못 받으셨을때만 가능하십니다.\n만약 밴, 등 이유로 환불을 해달라고 하시면 환불 안해드립니다.\n[ 모든 후원금은 전부 호스팅 비, 아니면 서버 패치비로 들어갑니다.]\n서명(인):\n구매 날짜: 0000년 0월 00일\n<이 서류를 작성하셔서 총관리자님께 후원문의를 하시면 됩니다.>")
    if message.content.startswith("!네이슈"):
        await message.channel.send("네이슈! 네이슈!")
    if message.content.startswith("!봇상태"):
        await message.channel.send("봇 상태: 양호\nping: 32ms\nHosting state: [신호 감지 불가]")

    if message.content.startswith("1시간"):
        a = datetime.datetime.today().year
        b = datetime.datetime.today().month
        c = datetime.datetime.today().day
        d = datetime.datetime.today().hour
        e = datetime.datetime.today().minute
        await client.channel.send(str(a) + "년 " + str(b) + "월 " + str(c) + "일 " + str(d) + "시 " + str(e) + "분 입니다.")

    if message.content.startswith("!1~3"):
        for x in range(3):
            await message.channel.send(x+1)

    if message.content.startswith("!주사위"):
        roll = message.content.split(" ")
        rolld = roll[1].split("d")
        dice = 0
        for i in range(0, int(rolld[0])+1):
            dice = dice + random.randint(1, int(rolld[1]))
        await message.channel.send(str(dice))

    if message.content.startswith("!골라"):
        choice = message.content.split(" ")
        choicenumber = random.randint(1, len(choice)-1)
        choiceresult = choice[choicenumber]
        await message.channel.send(choiceresult)

    if message.content.startswith("!뭐먹지"):
        food = "돈까스 탕수육 라면 밥 김치 굶기 육개장 카레 생선구이 김치찌게 된장찌게 짜장면 짬뽕 스파게티 피자 치킨 냉면 비빔냉면 가지볶음 전 회 계란후라이 계란말이 설렁탕 까르보나라 떡볶이 족발 보쌈 삼겹살 갈비 배추 쌈장 고추장 물 음료수 과자 시금치 간장"
        foodchoice = food.split(" ")
        foodnumber = random.randint(1, len(foodchoice))
        foodresult = foodchoice[foodnumber-1]
        await message.channel.send(foodresult)

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[1]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[2]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[3]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[4]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[5]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[6]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[7]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[8]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[9]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[10]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[11]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[12]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[13]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[14]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[15]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[16]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[17]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[18]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[19]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[20]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[21]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[22]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[23]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[24]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[25]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[26]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[27]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[28]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[29]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[30]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[31]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[32]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[33]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[34]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[35]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[36]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[37]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[38]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[39]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[40]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[41]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[42]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[43]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[44]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[45]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[46]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[47]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[48]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[49]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[50]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[51]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[52]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[53]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[54]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[55]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[56]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[57]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[58]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[59]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[60]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[61]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[62]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[63]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[64]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[65]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[66]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[67]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[68]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[69]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[70]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[71]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[72]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[73]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[74]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[75]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[76]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[77]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[78]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[79]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[80]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[81]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[82]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[83]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[84]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[85]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[86]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[87]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[88]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[89]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[90]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[91]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[92]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[93]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[94]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[95]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[96]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[97]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[98]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[99]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/사진"):
        pic = message.content.split(" ")[100]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("/채널메세지"):
        channel = message.content[7:25]
        msg = message.content[26:]
        await client.get_channel(int(channel)).send(msg)

    if message.content.startswith("/DM"):
        author = message.guild.get_member(int(message.content[4:22]))
        msg = message.content[23:]
        await author.send(msg)

    if message.content.startswith("/뮤트"):
        author = message.guild.get_member(int(message.content[4:22]))
        role = discord.utils.get(message.guild.roles, name="뮤트")
        await author.add_roles(role)

    if message.content.startswith("/언뮤트"):
        author = message.guild.get_member(int(message.content[5:23]))
        role = discord.utils.get(message.guild.roles, name="뮤트")
        await author.remove_roles(role)


    if message.content.startswith("/경고"):
        author = message.guild.get_member(int(message.content[4:22]))
        file = openpyxl.load_workbook("경고.xlsx")
        sheet = file.active
        i = 1
        while True:
            if sheet["A" + str(i)].value == str(message.author.id):
                sheet["B" + str(i)].value = int(sheet["B" + str(i)].value) + 1
                file.save("경고.xlsx")
                if sheet["B" + str(i)].value == 2:
                    await message.guild.ban(author)
                    await message.channel.send("경고 2회 누적입니다 서버에서 추방됩니다.")
                else:
                    await message.channel.send("경고 1회를 받았습니다")
                break
            if sheet["A" + str(i)].value == None:
                sheet["A" + str(i)].value = str(message.author.id)
                sheet["B" + str(i)].value = 1
                file.save("경고.xlsx")
                await message.channel.send("경고를 1회 받았습니다")
                break
            i += 1

    if "시발" in message.content and "병신" in message.content and "섹스" in message.content and "개새끼" in message.content and "개병신" in message.content and "쒸발롬아" in message.content and "시1발" in message.content and "ㅈ병신" in message.content and "개씨발련아" in message.content and "ㅈ같네" in message.content and "성관계" in message.content and "아 섹스하고 싶다" in message.content:
        file = openpyxl.load_workbook("경고.xlsx")
        sheet = flie.active
        member = discord.utils.get(client.get_all_members())
        for i in range(1, 31):
            if str(sheet["A" + str(i)].value) == str(message.author.id):
                sheet["B" + str(i)].value = int(sheet["B" + str(i)].value) + 1
                if int(sheet["B" + str(i)].value) == 3:
                    await client.ban(member, 1)
                break
                if str(sheet["A" + str(i)].value) == '-':
                    sheet["A" + str(i)].value = str(message.author.id)
                    sheet["B" + str(i)].value = 1
                    break
            file.save("경고.xlsx")
            await client.send_message(message.channel, "경고를 받으셨습니다. 3번을 받으시면 강퇴되니, 욕설, 섹드립, 등 하지마시길 바랍니다.")


    if message.content.startswith(""):
        file = openpyxl.load_workbook("레벨.xlsx")
        sheet = file.active
        exp = [100, 200, 300, 400, 500, 680, 890, 1029, 1265, 1598, 1972, 2085, 2439, 2598, 2892, 3192, 3201, 3923, 4012, 4293, 4683, 4872, 5016, 5287, 5467, 5698, 5982, 6098, 6293, 6432]
        i = 1
        while True:
            if sheet["A" + str(i)].value == str(message.author.id):
                sheet["B" + str(i)].value = sheet["B" + str(i)].value + 2
                if sheet["B" + str(i)].value >= exp[sheet["C" + str(i)].value]:
                    sheet["C" + str(i)].value = sheet["C" + str(i)].value + 1
                    await message.channel.send("레벨이 올랐습니다.\n현재 레벨 : " + str(sheet["C" + str(i)].value) + "\n경험치 : " + str(sheet["B" + str(i)].value))
                file.save("레벨.xlsx")
                break

            if sheet["A" + str(i)].value == None:
                sheet["A" + str(i)].value = str(message.author.id)
                sheet["B" + str(i)].value = 0
                sheet["C" + str(i)].value = 1
                file.save("레벨.xlsx")
                break

            i += 1

    if message.content.startswith("!역할설정"):
        role = ""
        rolename = message.content.split(" ")
        member = discord.utils.get(client.get_all_members(), id=rolename[1])
        for i in message.server.roles:
            if i.name == rolename[2]:
                role = i
                break
        await client.add_roles(member, role)

@client.event
async def on_reaction_add(reaction, user):
    if str(reaction.emoji) == "👍":
        await reaction.message.channel.send(user.name + "님이 엄지 척! 리액션을 하셨습니다.")
    if str(reaction.emoji) == "✔":
        await reaction.message.channel.send(user.name + "님이 확인을 하셨습니다.")
    if str(reaction.emoji) == "👏":
        await reaction.message.channel.send(user.name + "님이 축하를 해주셨습니다.")
    if str(reaction.emoji) == "✅":
        await reaction.message.channel.send(user.name + "님이 체크(확인)을 하셨습니다.")

i = 0
while True:
    i = i + 1
    if i == 2:
        print("Kidong Bot 작동 준비중 2초만 기다려주세요..")
        continue
    if i == 5:
        print("Kidong Bot Seting 중..")
        break
    print(i)

print("Kidong Bot Start!")


client.run("NjIyMDM1NzA0NjkxODE4NTE2.XavsVg.M3TGB7sMSWN-fEaQ6xXaawDW2kc")
